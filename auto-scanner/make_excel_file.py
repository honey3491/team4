import argparse
import json
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def load_json(path: Path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def sanitize_url_to_path(url: str | None) -> str:
    raw = str(url or "").strip()
    if not raw:
        return ""

    parsed = urlparse(raw)
    path = parsed.path or ""

    if not path:
        path = raw.split("?", 1)[0].split("#", 1)[0].strip()

    return path or "/"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        default="outputs/gpt_analysis_results.json",
        help="분석 결과 JSON 경로"
    )
    parser.add_argument(
        "--output",
        default="outputs/자동진단.xlsx",
        help="생성할 Excel 파일 경로"
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    data = load_json(input_path)

    # Workbook 생성
    wb = Workbook()

    # =========================
    # Metadata Sheet
    # =========================
    ws_meta = wb.active
    ws_meta.title = "metadata"

    meta_rows = [
        ["key", "value"],
        ["scan_id", data.get("scan_id")],
        ["target", data.get("target")],
        ["scanner", data.get("scanner")],
        ["generated_at", data.get("generated_at")],
        ["total", data["summary"]["total"]],
    ]

    severity = data["summary"]["severity"]

    for level in ["critical", "high", "medium", "low", "pass", "n/a", "pending"]:
        meta_rows.append([level, severity.get(level, 0)])

    for row in meta_rows:
        ws_meta.append(row)

    # 헤더 스타일
    for cell in ws_meta[1]:
        cell.font = Font(bold=True)

    # 컬럼 너비 자동 조정
    for col in ws_meta.columns:
        max_length = 0
        column = get_column_letter(col[0].column)

        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws_meta.column_dimensions[column].width = max_length + 5

    # =========================
    # Findings Sheet
    # =========================
    ws_find = wb.create_sheet(title="findings")

    headers = [
        "scan_id",
        "target",
        "generated_at",
        "id",
        "check_id",
        "owasp",
        "vuln_name",
        "url",
        "severity",
        "evidence",
        "recommendation"
    ]

    ws_find.append(headers)

    for finding in data["results"]:
        row = [
            data.get("scan_id"),
            data.get("target"),
            data.get("generated_at"),
            finding.get("id"),
            finding.get("check_id"),
            finding.get("owasp"),
            finding.get("name"),
            sanitize_url_to_path(finding.get("url")),
            finding.get("severity"),
            finding.get("evidence"),
            finding.get("recommendation")
        ]

        ws_find.append(row)

    # 헤더 스타일
    for cell in ws_find[1]:
        cell.font = Font(bold=True)

    # 컬럼 너비 자동 조정
    for col in ws_find.columns:
        max_length = 0
        column = get_column_letter(col[0].column)

        for cell in col:
            try:
                value = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(value[:100]))
            except Exception:
                pass

        ws_find.column_dimensions[column].width = min(max_length + 5, 80)

    wb.save(output_path)
    print(f"[+] Excel 파일 저장 완료: {output_path}")


if __name__ == "__main__":
    main()
